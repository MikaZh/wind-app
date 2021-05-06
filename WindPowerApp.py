# -*- coding: utf-8 -*-

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt
import sys
import pyowm
import openpyxl
from PyQt5.QtWidgets import QMessageBox, QWidget, QToolTip, QPushButton, QApplication, QLabel
from PyQt5.QtGui import QFont, QMovie
from forall import Ui_Form
from about import Ui_about
from datetime import date, datetime

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(654, 578)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("main.png"), QtGui.QIcon.Normal, QtGui.QIcon.On)
        MainWindow.setWindowIcon(icon)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(150, 30, 391, 71))
        font = QtGui.QFont()
        font.setPointSize(28)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(40, 120, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(40, 190, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(40, 330, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(170, 120, 371, 41))
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("---SELECT THE REGION---")
        self.comboBox.addItem("---FOR ALL REGIONS---")
        self.comboBox.addItem("Akmola Region")
        self.comboBox.addItem("Aktobe Region")
        self.comboBox.addItem("Almaty")
        self.comboBox.addItem("Almaty Region")
        self.comboBox.addItem("Atyrau Region")
        self.comboBox.addItem("East Kazakhstan Region")
        self.comboBox.addItem("Jambyl Region")
        self.comboBox.addItem("Karagandy Region")
        self.comboBox.addItem("Kostanay Region")
        self.comboBox.addItem("Kyzylorda Region")
        self.comboBox.addItem("Mangystau Region")
        self.comboBox.addItem("North Kazakhstan Region")
        self.comboBox.addItem("Nur-Sultan")
        self.comboBox.addItem("Pavlodar Region")
        self.comboBox.addItem("Shymkent")
        self.comboBox.addItem("Turkistan Region")
        self.comboBox.addItem("West Kazakhstan Region")

        self.areaLabel = QtWidgets.QLineEdit(self.centralwidget)
        self.areaLabel.setGeometry(QtCore.QRect(170, 190, 371, 41))
        self.areaLabel.setAcceptDrops(True)
        self.areaLabel.setAutoFillBackground(False)
        self.areaLabel.setInputMethodHints(QtCore.Qt.ImhPreferNumbers)
        self.areaLabel.setInputMask("")
        self.areaLabel.setReadOnly(True)
        self.areaLabel.setPlaceholderText("")
        self.areaLabel.setStyleSheet("color: grey;")
        self.areaLabel.setClearButtonEnabled(True)
        self.areaLabel.setObjectName("areaLabel")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(170, 330, 371, 41))
        self.lineEdit_2.setInputMethodHints(QtCore.Qt.ImhPreferNumbers)
        self.lineEdit_2.setClearButtonEnabled(True)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(230, 470, 181, 41))
        self.pushButton.setObjectName("pushButton")
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setGeometry(QtCore.QRect(550, 190, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        self.label_6.setGeometry(QtCore.QRect(550, 330, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(self.centralwidget)
        self.label_7.setGeometry(QtCore.QRect(550, 260, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.powerLabel= QtWidgets.QLabel(self.centralwidget)
        self.powerLabel.setObjectName("powerLabel")
        self.powerLabel.setGeometry(QtCore.QRect(610,320,31,41))
        font2 = QFont()
        font2.setPointSize(12)
        self.powerLabel.setFont(font2)
        self.label_8 = QtWidgets.QLabel(self.centralwidget)
        self.label_8.setGeometry(QtCore.QRect(40, 260, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.areaLabel_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.areaLabel_2.setGeometry(QtCore.QRect(170, 260, 371, 41))
        self.areaLabel_2.setAcceptDrops(True)
        self.areaLabel_2.setAutoFillBackground(False)
        self.areaLabel_2.setInputMethodHints(QtCore.Qt.ImhPreferNumbers)
        self.areaLabel_2.setInputMask("")
        self.areaLabel_2.setText("")
        self.areaLabel_2.setPlaceholderText("")
        self.areaLabel_2.setClearButtonEnabled(True)
        self.areaLabel_2.setObjectName("areaLabel_2")
        self.pushButton_3 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_3.setGeometry(QtCore.QRect(10, 10, 41, 41))
        self.pushButton_3.setText("")
        self.pushButton_3.setToolTip('About the program')
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("about.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_3.setIcon(icon1)
        self.pushButton_3.setIconSize(QtCore.QSize(25, 25))
        self.pushButton_3.setObjectName("pushButton_3")
        self.label_9 = QtWidgets.QLabel(self.centralwidget)
        self.label_9.setGeometry(QtCore.QRect(550, 400, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_9.setFont(font)
        self.label_9.setObjectName("label_9")
        self.label_10 = QtWidgets.QLabel(self.centralwidget)
        self.label_10.setGeometry(QtCore.QRect(40, 400, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_10.setFont(font)
        self.label_10.setObjectName("label_10")
        self.areaLabel_3 = QtWidgets.QLineEdit(self.centralwidget)
        self.areaLabel_3.setGeometry(QtCore.QRect(170, 400, 371, 41))
        self.areaLabel_3.setAcceptDrops(True)
        self.areaLabel_3.setAutoFillBackground(False)
        self.areaLabel_3.setInputMethodHints(QtCore.Qt.ImhPreferNumbers)
        self.areaLabel_3.setInputMask("")
        self.areaLabel_3.setText("")
        self.areaLabel_3.setPlaceholderText("")
        self.areaLabel_3.setClearButtonEnabled(True)
        self.areaLabel_3.setObjectName("areaLabel_3")
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionOne_Region = QtWidgets.QAction(MainWindow)
        self.actionOne_Region.setObjectName("actionOne_Region")
        self.actionAll_Regions = QtWidgets.QAction(MainWindow)
        self.actionAll_Regions.setObjectName("actionAll_Regions")
        self.actionLast_update = QtWidgets.QAction(MainWindow)
        self.actionLast_update.setObjectName("actionLast_update")

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Wind Power Calculator"))
        self.label.setText(_translate("MainWindow", "Power Calculation"))
        self.label_2.setText(_translate("MainWindow", "Region"))
        self.label_3.setText(_translate("MainWindow", "Speed"))
        self.label_4.setText(_translate("MainWindow", "Air density"))
        self.areaLabel.setText(_translate("MainWindow", "No need to fill: data taken online"))
        self.lineEdit_2.setText(_translate("MainWindow", "1.24"))
        self.pushButton.setText(_translate("MainWindow", "Calculate"))
        self.pushButton.clicked.connect(self.calculate)
        self.pushButton_3.clicked.connect(self.about)
        self.label_9.setText(_translate("MainWindow", "%"))
        self.label_10.setText(_translate("MainWindow", "Efficiency"))
        #self.pushButton_5.clicked.connect(self.openForm)
        self.label_5.setText(_translate("MainWindow", "m/s"))
        self.label_6.setText(_translate("MainWindow", "kg/m"))
        self.label_7.setText(_translate("MainWindow", "m"))
        self.label_8.setText(_translate("MainWindow", "Radius"))
        self.cities = ["Kokshetau", "Aktobe", "Almaty", "Taldykorgan", "Atyrau", "Oskemen", "Taraz", "Karagandy", "Kostanay", "Kyzylorda", "Aktau", "Petropavl", "Nur-Sultan", "Pavlodar", "Shymkent", "Turkistan", "Oral"]
        self.powerLabel.setText(_translate("MainWindow", "3"))
    def calculate(self):
        owm = pyowm.OWM('a4059161b180b05d8caa7275375bdf1d')
        mgr = owm.weather_manager()
        self.region = str(self.comboBox.currentText())
        self.area = self.areaLabel_2.text()
        self.density = self.lineEdit_2.text()
        self.cp = self.areaLabel_3.text()
        try:
            number = float(self.area)
            number2 = float(self.cp)
            city = str(self.reg_select(self.region))
            if((city!="default") and (city!="all")):
                loc = mgr.weather_at_place(city)
                weather = loc.weather
                wind = weather.wind()
                self.speed = wind.get('speed')
                self.areaLabel.setText(str(self.speed))
                self.power = (0.5*(float(self.cp)/100)*float(self.density)*3.1415*(float(self.area)**2)*(self.speed**3))/1000
                self.label.setText("Power is: "+str(round(self.power, 3))+" kW")
                self.update()
                self.history()

            elif(city=="default"):
                d = QMessageBox()
                d.setWindowTitle("Region is not selected")
                d.setText("Please, select the region from the dropdown menu")
                d.setIcon(QMessageBox.Warning)
                x = d.exec_()
                pass
            elif(city=="all"):
                self.update()
                self.total = []
                self.sum = 0
                for each in self.cities:
                    self.region = each
                    loc = mgr.weather_at_place(each)
                    weather = loc.weather
                    wind = weather.wind()
                    self.speed = wind.get('speed')
                    #self.areaLabel.setText(str(speed))
                    self.power = (0.5*(float(self.cp)/100)*float(self.density)*3.1415*(float(self.area)**2)*(self.speed**3))/1000
                    self.sum += self.power
                    self.total.append(str(round(self.power,3))+" kW")
                    self.history()
                self.total.append(str(round(self.sum,3))+" kW")
                self.openForm()

        except Exception:
            msg = QMessageBox()
            msg.setWindowTitle("Wrong Input")
            msg.setText("Invalid Input type. Please, enter the values properly.")
            msg.setIcon(QMessageBox.Warning)
            msg.setDetailedText("- Decimal point has to be seperated by the dot(.), not comma (,). \n - Area, Density and Efficiency boxes cannot be empty. \n -Efficiency is in percents (1-100%)")
            x = msg.exec_()
            pass

    def about(self):
        self.window = QtWidgets.QMainWindow()
        self.ui = Ui_about()
        self.ui.setupUi(self.window)
        self.window.show()

    def update(self):
        self.label.adjustSize()

    def history(self):
        filename = 'history.xlsx'
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        last_row = ws.max_row+1
        #today = date.today()
        #date_now = today.strftime("%b-%d-%Y")
        time_now = datetime.now()

        #ws.cell(column=1, row=last_row, value=date_now)
        ws.cell(column=1, row=last_row, value=time_now)
        ws.cell(column=2, row=last_row, value=self.region)
        ws.cell(column=3, row=last_row, value=self.speed)
        ws.cell(column=4, row=last_row, value=float(self.area))
        ws.cell(column=5, row=last_row, value=float(self.density))
        ws.cell(column=6, row=last_row, value=round(self.power,3))
        wb.save(filename)

    def openForm(self):
        self.window = QtWidgets.QMainWindow()
        self.message = self.total
        self.ui = Ui_Form(self.message)
        self.ui.setupUi(self.window)
        self.window.show()

    def reg_select(self, reg):
        if(reg=="Aktobe Region"):
            return "Aktobe"
        elif(reg=="---SELECT THE REGION---"):
            return "default"
        elif(reg=="---FOR ALL REGIONS---"):
            return "all"
        elif(reg=="Akmola Region"):
            return "Kokshetau"
        elif(reg=="Almaty"):
            return "Almaty"
        elif(reg=="Almaty Region"):
            return "Taldykorgan"
        elif(reg=="Atyrau Region"):
            return "Atyrau"
        elif(reg=="East Kazakhstan Region"):
            return "Oskemen"
        elif(reg=="Jambyl Region"):
            return "Taraz"
        elif(reg=="Karagandy Region"):
            return "Karagandy"
        elif(reg=="Kostanay Region"):
            return"Kostanay"
        elif(reg=="Kyzylorda Region"):
            return "Kyzylorda"
        elif(reg=="Mangystau Region"):
            return "Aktau"
        elif(reg=="North Kazakhstan Region"):
            return "Petropavl"
        elif(reg=="Nur-Sultan"):
            return "Nur-Sultan"
        elif(reg=="Pavlodar Region"):
            return "Pavlodar"
        elif(reg=="Shymkent"):
            return "Shymkent"
        elif(reg=="Turkistan Region"):
            return "Turkistan"
        elif(reg=="West Kazakhstan Region"):
            return "Oral"

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
